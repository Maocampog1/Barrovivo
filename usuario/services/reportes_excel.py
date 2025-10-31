from typing import Iterable, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .reportes_interface import ReporteVentasGenerator

class ExcelReporteVentasGenerator(ReporteVentasGenerator):
    filename = "reporte_ventas.xlsx"

    def render(self, pedidos: Iterable[Any]) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        headers = [
            "ID Pedido", "Fecha", "Cliente", "Correo",
            "Departamento", "Municipio", "Dirección", "Total",
            "ID Item", "Producto", "Cantidad", "Precio"
        ]
        ws.append(headers)

        for p in pedidos:
            items = list(p.items.all()) or [None]
            for it in items:
                ws.append([
                    p.id, p.fecha.strftime("%Y-%m-%d %H:%M"),
                    p.nombre_cliente, p.correo,
                    p.departamento, p.municipio, p.direccion,
                    float(p.total),
                    (it.id if it else ""),
                    (it.producto.nombre if it else ""),
                    (it.cantidad if it else ""),
                    (float(it.precio) if it else ""),
                ])

        for col in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
